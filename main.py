import datetime

import openpyxl
from PyQt6.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook
from selenium import webdriver


class TaskThread(QThread):
    callback = pyqtSignal(object)

    def __init__(self, func):
        super(TaskThread, self).__init__()
        self.func = func

    def run(self):
        self.callback.emit(self.func())


class Checker:

    def __init__(self):

        f = open("exe.js", "r")
        self.script = f.read()

        self.result_book = openpyxl.Workbook()
        self.result_sheet = self.result_book.worksheets[0]
        self.__stop = False

    def stop(self):
        self.__stop = True

    def open_account(self, file_path, on_finish):
        self.thread = TaskThread(lambda: self.__open_account(file_path))
        self.thread.callback.connect(on_finish)
        self.thread.start()

    def __open_account(self, file_path):
        self.save_path = f'{datetime.datetime.now().strftime("%y%m%d%H%M%S")}_result.xlsx'
        driver = webdriver.Chrome()

        workbook = load_workbook(file_path)
        sheet = workbook.worksheets[0]

        accounts = {}
        for row in sheet.rows:
            accounts[row[0].value] = row[1].value

        for key in accounts.keys():
            if self.__stop:
                break
            self.__open_web(driver, key, accounts[key])

        self.result_book.save(self.save_path)
        return f'文件已保存->{self.save_path}'

    def __open_web(self, driver, account_no, account_email):

        driver.get("https://www.adidas.com/us/order-tracker")
        driver.execute_script(self.script, account_no, account_email)

        info = ''
        loaded = True
        while (loaded):
            try:
                info = self.get_info(driver, account_no, account_email)
                loaded = False
            except:
                loaded = True
        return info

    def check_get(self, driver, js):
        try:
            return driver.execute_script(js)
        except:
            return "无"

    def get_info(self, driver, account_no, account_email):
        load = driver.execute_script(
            "return window.document.getElementsByClassName('img_with_fallback___2aHBu')[0].src;")
        info = [
            account_no,
            account_email,
            self.check_get(driver,
                           "return window.document.getElementsByClassName('label___1o2Is')[0].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('tracking-description___3iTmt')[0].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('product-card__attributes--name___2dtQ3')[0].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('product-card__attributes--price___1YdLZ')[0].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('product-card__attributes--sec___3nKjc')[0].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('no-left-padding___27BwT gl-vspace-bpall-small col-s-12 col-m-12 col-l-24')[0].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('no-left-padding___27BwT gl-vspace-bpall-small col-s-12 col-m-6 col-l-12')[0].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('address___3gN3A')[0].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('address___3gN3A')[1].textContent;"),
            self.check_get(driver,
                           "return window.document.getElementsByClassName('gl-vspace-bpall-medium col-s-12 col-m-6 col-l-12 no-gutters no-left-padding___BSrES')[0].textContent;")
        ]
        self.result_sheet.append(info)
        self.result_book.save(self.save_path)
        return info

    if __name__ == '__main__':
        print('init')
